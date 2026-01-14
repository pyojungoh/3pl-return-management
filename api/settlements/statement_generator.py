"""
정산 명세서 엑셀 생성 모듈
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


# 회사 정보
COMPANY_INFO = {
    'name': '제이제이',
    'business_number': '803-15-01907',
    'representative': '이정민',
    'address': '경기도 김포시 통진읍 월하로 352-18',
    'business_type': '업태: 도매 및 소매업, 종목: 상품 종합 도매업',
    'opening_date': '2022년 03월 04일'
}

# 입금 계좌 정보
BANK_INFO = {
    'bank': 'KB국민은행',
    'account_number': '910601-01-494700',
    'account_holder': '이정민(제이제이)'
}


def create_settlement_statement(settlement_data):
    """
    정산 명세서 엑셀 파일 생성
    
    Args:
        settlement_data: 정산 데이터 딕셔너리
    
    Returns:
        BytesIO: 엑셀 파일 바이너리 데이터
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "정산명세서"
    
    # 스타일 정의
    header_font = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')
    title_font = Font(name='맑은 고딕', size=16, bold=True)
    label_font = Font(name='맑은 고딕', size=11, bold=True)
    content_font = Font(name='맑은 고딕', size=11)
    number_font = Font(name='맑은 고딕', size=11)
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    row = 1
    
    # 제목
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = '정산 명세서'
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = title_fill
    row += 2
    
    # 회사 정보 섹션
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = '■ 공급자 정보'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].alignment = left_align
    row += 1
    
    ws[f'A{row}'] = '상호'
    ws[f'B{row}'] = COMPANY_INFO['name']
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].font = content_font
    ws[f'A{row}'].border = border_style
    ws[f'B{row}'].border = border_style
    row += 1
    
    ws[f'A{row}'] = '사업자등록번호'
    ws[f'B{row}'] = COMPANY_INFO['business_number']
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].font = content_font
    ws[f'A{row}'].border = border_style
    ws[f'B{row}'].border = border_style
    row += 1
    
    ws[f'A{row}'] = '대표자'
    ws[f'B{row}'] = COMPANY_INFO['representative']
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].font = content_font
    ws[f'A{row}'].border = border_style
    ws[f'B{row}'].border = border_style
    row += 1
    
    ws[f'A{row}'] = '사업장소재지'
    ws[f'B{row}'] = COMPANY_INFO['address']
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].font = content_font
    ws[f'A{row}'].border = border_style
    ws[f'B{row}'].border = border_style
    row += 2
    
    # 공급받는자 정보
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = '■ 공급받는자 정보'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].alignment = left_align
    row += 1
    
    ws[f'A{row}'] = '상호'
    ws[f'B{row}'] = settlement_data.get('company_name', '')
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].font = content_font
    ws[f'A{row}'].border = border_style
    ws[f'B{row}'].border = border_style
    row += 1
    
    ws[f'A{row}'] = '정산년월'
    ws[f'B{row}'] = settlement_data.get('settlement_year_month', '')
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].font = content_font
    ws[f'A{row}'].border = border_style
    ws[f'B{row}'].border = border_style
    row += 2
    
    # 정산 내역 테이블 헤더
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = '■ 정산 내역'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].alignment = left_align
    row += 1
    
    headers = ['항목', '금액 (원)']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = border_style
    row += 1
    
    # 정산 항목
    work_fee = settlement_data.get('work_fee', 0) or 0
    inout_fee = settlement_data.get('inout_fee', 0) or 0
    shipping_fee = settlement_data.get('shipping_fee', 0) or 0
    storage_fee = settlement_data.get('storage_fee', 0) or 0
    special_work_fee = settlement_data.get('special_work_fee', 0) or 0
    error_deduction = settlement_data.get('error_deduction', 0) or 0
    
    items = [
        ('작업비', work_fee),
        ('입출고비', inout_fee),
        ('택배비', shipping_fee),
        ('보관료', storage_fee),
        ('특수작업', special_work_fee),
        ('오배송/누락 차감', -error_deduction),
    ]
    
    for item_name, amount in items:
        ws.cell(row=row, column=1, value=item_name).font = content_font
        ws.cell(row=row, column=1).alignment = left_align
        ws.cell(row=row, column=1).border = border_style
        
        amount_cell = ws.cell(row=row, column=2, value=amount)
        amount_cell.font = number_font
        amount_cell.alignment = right_align
        amount_cell.border = border_style
        amount_cell.number_format = '#,##0'
        row += 1
    
    # 총액
    total_amount = work_fee + inout_fee + shipping_fee + storage_fee + special_work_fee - error_deduction
    
    ws.cell(row=row, column=1, value='총액').font = label_font
    ws.cell(row=row, column=1).alignment = left_align
    ws.cell(row=row, column=1).border = border_style
    ws.cell(row=row, column=1).fill = title_fill
    
    total_cell = ws.cell(row=row, column=2, value=total_amount)
    total_cell.font = Font(name='맑은 고딕', size=11, bold=True)
    total_cell.alignment = right_align
    total_cell.border = border_style
    total_cell.fill = title_fill
    total_cell.number_format = '#,##0'
    row += 1
    
    # 부가세
    tax = int(total_amount * 0.1)
    ws.cell(row=row, column=1, value='부가세 (10%)').font = label_font
    ws.cell(row=row, column=1).alignment = left_align
    ws.cell(row=row, column=1).border = border_style
    
    tax_cell = ws.cell(row=row, column=2, value=tax)
    tax_cell.font = number_font
    tax_cell.alignment = right_align
    tax_cell.border = border_style
    tax_cell.number_format = '#,##0'
    row += 1
    
    # 결제금액
    final_amount = total_amount + tax
    ws.cell(row=row, column=1, value='결제금액').font = label_font
    ws.cell(row=row, column=1).alignment = left_align
    ws.cell(row=row, column=1).border = border_style
    ws.cell(row=row, column=1).fill = title_fill
    
    final_cell = ws.cell(row=row, column=2, value=final_amount)
    final_cell.font = Font(name='맑은 고딕', size=12, bold=True)
    final_cell.alignment = right_align
    final_cell.border = border_style
    final_cell.fill = title_fill
    final_cell.number_format = '#,##0'
    row += 2
    
    # 입금 계좌 정보
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = '■ 입금 계좌 정보'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].alignment = left_align
    row += 1
    
    ws[f'A{row}'] = '은행'
    ws[f'B{row}'] = BANK_INFO['bank']
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].font = content_font
    ws[f'A{row}'].border = border_style
    ws[f'B{row}'].border = border_style
    row += 1
    
    ws[f'A{row}'] = '계좌번호'
    ws[f'B{row}'] = BANK_INFO['account_number']
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].font = content_font
    ws[f'A{row}'].border = border_style
    ws[f'B{row}'].border = border_style
    row += 1
    
    ws[f'A{row}'] = '예금주'
    ws[f'B{row}'] = BANK_INFO['account_holder']
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].font = content_font
    ws[f'A{row}'].border = border_style
    ws[f'B{row}'].border = border_style
    row += 2
    
    # 비고
    if settlement_data.get('memo'):
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = '■ 비고'
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].alignment = left_align
        row += 1
        
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = settlement_data.get('memo', '')
        ws[f'A{row}'].font = content_font
        ws[f'A{row}'].alignment = left_align
        ws[f'A{row}'].border = border_style
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # 파일을 BytesIO로 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
