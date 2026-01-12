"""
구글 드라이브 파일 업로드 테스트 스크립트
"""
import os
import sys
import tempfile
from datetime import datetime

# 프로젝트 루트를 Python 경로에 추가
sys.path.insert(0, os.path.dirname(__file__))

def create_test_excel_file():
    """테스트용 엑셀 파일 생성"""
    try:
        import openpyxl
        
        # 임시 파일 생성
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file_path = temp_file.name
        temp_file.close()
        
        # 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "테스트"
        
        # 테스트 데이터 추가
        ws['A1'] = '테스트'
        ws['B1'] = '데이터'
        ws['A2'] = '1'
        ws['B2'] = '2'
        
        # 파일 저장
        wb.save(temp_file_path)
        
        # 파일 데이터 읽기
        with open(temp_file_path, 'rb') as f:
            file_data = f.read()
        
        # 임시 파일 삭제
        os.unlink(temp_file_path)
        
        return file_data, 'test_upload_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.xlsx'
    except ImportError:
        print("⚠️ openpyxl이 설치되지 않았습니다. 간단한 바이너리 파일로 테스트합니다.")
        # 간단한 바이너리 데이터로 테스트
        test_data = b'Test Excel File Content'
        return test_data, 'test_upload_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.xlsx'
    except Exception as e:
        print(f"❌ 테스트 파일 생성 실패: {e}")
        return None, None

def test_excel_upload():
    """엑셀 파일 업로드 테스트"""
    print("=" * 60)
    print("구글 드라이브 엑셀 파일 업로드 테스트")
    print("=" * 60 + "\n")
    
    try:
        from api.uploads.google_drive import upload_excel_to_drive
        
        # 테스트 파일 생성
        print("1. 테스트 파일 생성 중...")
        file_data, filename = create_test_excel_file()
        if not file_data:
            print("❌ 테스트 파일 생성 실패")
            return False
        
        print(f"✅ 테스트 파일 생성 성공: {filename} ({len(file_data)} bytes)")
        
        # 파일 업로드
        print(f"\n2. 파일 업로드 중...")
        print(f"   파일명: {filename}")
        print(f"   파일 크기: {len(file_data)} bytes")
        print(f"   대상 폴더: 정산파일")
        
        result = upload_excel_to_drive(
            file_data=file_data,
            filename=filename,
            folder_name='정산파일'
        )
        
        if result.get('success'):
            print(f"\n✅ 파일 업로드 성공!")
            print(f"   파일 ID: {result.get('file_id')}")
            print(f"   파일 URL: {result.get('file_url')}")
            print(f"   웹 보기 링크: {result.get('web_view_link')}")
            print(f"   메시지: {result.get('message')}")
            return True
        else:
            print(f"\n❌ 파일 업로드 실패")
            print(f"   메시지: {result.get('message')}")
            return False
            
    except Exception as e:
        print(f"\n❌ 테스트 오류: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """메인 테스트 함수"""
    print("\n" + "=" * 60)
    print("구글 드라이브 파일 업로드 테스트 시작")
    print("=" * 60 + "\n")
    
    success = test_excel_upload()
    
    print("\n" + "=" * 60)
    print("테스트 결과")
    print("=" * 60)
    if success:
        print("✅ 파일 업로드 테스트 통과!")
        print("구글 드라이브 API 연동이 정상적으로 작동합니다.")
    else:
        print("❌ 파일 업로드 테스트 실패")
        print("위의 오류 메시지를 확인하세요.")
    print("=" * 60 + "\n")

if __name__ == '__main__':
    main()

